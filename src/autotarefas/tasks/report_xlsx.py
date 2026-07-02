"""
Geracao do artefato profissional da Auditoria: ``planilha_validada.xlsx``.

Monta, com openpyxl, uma pasta de trabalho de 4 abas pensada para o
cliente abrir no Excel:

- **Resumo**            — cabecalho, contadores (total/validos/invalidos/
  normalizados) e a tabela de erros por categoria.
- **Registros validos** — as linhas validas (ja normalizadas no modo
  limpeza), com cabecalho formatado, autofiltro e painel congelado.
- **Registros invalidos** — as linhas invalidas + coluna ``motivo``.
- **Auditoria**         — o audit trail (antes/depois/regras) do modo
  limpeza; nos demais modos, uma nota de que nada foi normalizado.

E um RELATORIO (retrato do resultado), nao um modelo editavel: os
contadores sao gravados como valores (fatos da auditoria), sem formulas.
Nao inventa dado — apenas organiza e apresenta o que a validacao produziu.
"""

from __future__ import annotations

import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from autotarefas.core.base import TaskResult
from autotarefas.tasks.artifacts import REASON_COLUMN, split_valid_invalid

#: Nome fixo do artefato XLSX.
XLSX_NAME = "planilha_validada.xlsx"

# ============================================================
# Paleta e estilos (fonte profissional, cores por status)
# ============================================================

_FONT = "Arial"
_TITLE_FONT = Font(name=_FONT, bold=True, size=16, color="1F4E78")
_LABEL_FONT = Font(name=_FONT, bold=True)
_HEADER_FONT = Font(name=_FONT, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")  # azul escuro
_OK_FILL = PatternFill("solid", fgColor="E2EFDA")  # verde claro
_ERROR_FILL = PatternFill("solid", fgColor="FCE4E4")  # vermelho claro
_NEUTRAL_FILL = PatternFill("solid", fgColor="FFF2CC")  # amarelo claro
_CENTER = Alignment(horizontal="center", vertical="center")

_SHEET_RESUMO = "Resumo"
_SHEET_VALIDOS = "Registros validos"
_SHEET_INVALIDOS = "Registros invalidos"
_SHEET_AUDITORIA = "Auditoria"

#: Rotulos legiveis para as categorias de erro (aba Resumo).
_CATEGORY_LABELS = {
    "cpf": "CPF invalido",
    "cnpj": "CNPJ invalido",
    "email": "E-mail invalido",
    "telefone": "Telefone invalido",
    "obrigatorio": "Campo obrigatorio vazio",
    "duplicado": "Duplicados",
    "tamanho": "Texto muito curto",
    "intervalo": "Fora do intervalo",
    "enum": "Valor nao permitido",
    "tipo": "Tipo invalido",
    "formato": "Formato invalido",
    "outro": "Outros",
}


# ============================================================
# Helpers
# ============================================================


def _cell_value(value: object) -> object:
    """Converte celula do DataFrame para um tipo que o openpyxl aceita."""
    # None ou NaN viram string vazia.
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    if isinstance(value, (bool, int, float, str)):
        return value
    # numpy int64/float64 e afins: converte para nativo via item().
    item = getattr(value, "item", None)
    if callable(item):
        native = item()
        if isinstance(native, (bool, int, float, str)):
            return native
    return str(value)


def _style_header_row(ws: Worksheet, n_cols: int) -> None:
    """Aplica estilo ao cabecalho (linha 1), congela painel e liga o filtro."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
    ws.freeze_panes = "A2"
    last_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def _autofit_columns(ws: Worksheet, max_width: int = 60) -> None:
    """Ajusta a largura das colunas ao maior conteudo (com um teto)."""
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _write_dataframe(ws: Worksheet, dataframe: pd.DataFrame) -> None:
    """Escreve cabecalho + linhas do DataFrame e aplica formatacao padrao."""
    ws.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        ws.append([_cell_value(v) for v in row])
    data_font = Font(name=_FONT)
    for cell_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in cell_row:
            cell.font = data_font
    _style_header_row(ws, len(dataframe.columns))
    _autofit_columns(ws)


def _lines_to_indices(lines: list[int]) -> list[int]:
    return [n - 2 for n in lines]


def _rules_str(rules: object) -> str:
    """Junta as regras de uma alteracao em texto (lida com tipagem object)."""
    if isinstance(rules, list):
        return ", ".join(str(r) for r in rules)
    return ""


# ============================================================
# Abas
# ============================================================


def _build_resumo(ws: Worksheet, result: TaskResult, n_invalid: int, n_valid: int) -> None:
    data = result.data
    ws["A1"] = "Auditoria de planilha"
    ws["A1"].font = _TITLE_FONT

    rows: list[tuple[str, object]] = [
        ("Arquivo", str(data.get("file", ""))),
        ("Modo", str(data.get("mode", ""))),
        ("Total de registros", int(data.get("rows", 0))),
        ("Registros validos", n_valid),
        ("Registros invalidos", n_invalid),
        ("Valores normalizados", int(data.get("total_cleaned", 0))),
    ]
    fills = {
        "Registros validos": _OK_FILL,
        "Registros invalidos": _ERROR_FILL,
        "Valores normalizados": _NEUTRAL_FILL,
    }
    start = 3
    for offset, (label, value) in enumerate(rows):
        r = start + offset
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = Font(name=_FONT)
        if label in fills:
            ws.cell(row=r, column=1).fill = fills[label]
            cell.fill = fills[label]

    # Tabela de erros por categoria
    header_row = start + len(rows) + 1
    ws.cell(row=header_row, column=1, value="Erros por categoria").font = _LABEL_FONT
    ws.cell(row=header_row + 1, column=1, value="Categoria").font = _HEADER_FONT
    ws.cell(row=header_row + 1, column=1).fill = _HEADER_FILL
    ws.cell(row=header_row + 1, column=2, value="Quantidade").font = _HEADER_FONT
    ws.cell(row=header_row + 1, column=2).fill = _HEADER_FILL

    by_category: dict[str, int] = data.get("issues_by_category", {})
    line = header_row + 2
    if by_category:
        for category, count in by_category.items():
            label = _CATEGORY_LABELS.get(category, category)
            ws.cell(row=line, column=1, value=label).font = Font(name=_FONT)
            ws.cell(row=line, column=2, value=count).font = Font(name=_FONT)
            line += 1
    else:
        ws.cell(row=line, column=1, value="Nenhum problema encontrado").font = Font(name=_FONT)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def _build_auditoria(ws: Worksheet, result: TaskResult) -> None:
    changes: list[dict[str, object]] = result.data.get("cleaning_changes", [])
    mode = str(result.data.get("mode", ""))

    if not changes:
        note = (
            "Nenhuma normalizacao aplicada."
            if mode == "limpeza"
            else f"Modo '{mode}': dados nao sao alterados (sem normalizacao)."
        )
        ws["A1"] = note
        ws["A1"].font = Font(name=_FONT)
        ws.column_dimensions["A"].width = 60
        return

    audit_df = pd.DataFrame(
        {
            "linha": [c.get("line") for c in changes],
            "coluna": [c.get("column") for c in changes],
            "antes": [c.get("before") for c in changes],
            "depois": [c.get("after") for c in changes],
            "regras": [_rules_str(c.get("rules")) for c in changes],
        }
    )
    _write_dataframe(ws, audit_df)


# ============================================================
# Entry point
# ============================================================


def write_xlsx_report(dataframe: pd.DataFrame, result: TaskResult, path: Path) -> None:
    """
    Gera ``planilha_validada.xlsx`` com as 4 abas.

    Args:
        dataframe: DataFrame processado (normalizado no modo limpeza).
        result: TaskResult da validacao.
        path: Caminho do arquivo .xlsx a criar.
    """
    valid_lines, invalid_lines, reasons = split_valid_invalid(result)

    wb = Workbook()
    resumo = wb.active
    resumo.title = _SHEET_RESUMO
    _build_resumo(resumo, result, n_invalid=len(invalid_lines), n_valid=len(valid_lines))

    # Registros validos
    validos_df = dataframe.iloc[_lines_to_indices(valid_lines)]
    _write_dataframe(wb.create_sheet(_SHEET_VALIDOS), validos_df)

    # Registros invalidos (+ motivo)
    invalidos_df = dataframe.iloc[_lines_to_indices(invalid_lines)].copy()
    invalidos_df[REASON_COLUMN] = [" | ".join(reasons.get(n, [])) for n in invalid_lines]
    _write_dataframe(wb.create_sheet(_SHEET_INVALIDOS), invalidos_df)

    # Auditoria
    _build_auditoria(wb.create_sheet(_SHEET_AUDITORIA), result)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


__all__ = ["XLSX_NAME", "write_xlsx_report"]
