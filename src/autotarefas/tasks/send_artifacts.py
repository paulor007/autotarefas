"""
Artefatos do Cadastro automatico via planilha (send_api --out-dir).

A partir do DataFrame lido e do TaskResult (que ja carrega os items
estruturados da fase 2/3), gera o pacote canonico de 4 artefatos —
espelho do contrato da Auditoria de planilha:

- ``importacao_report.json``    relatorio tecnico/maquina (contrato do front)
- ``registros_enviados.csv``    linhas que ENTRARAM + status + id externo
- ``registros_falhos.csv``      linhas que falharam + motivo/categoria —
                                REENVIAVEL por construcao (ver abaixo)
- ``importacao_resultado.xlsx`` planilha executiva de 4 abas

Reenviabilidade por construcao:
    As colunas de metadado dos CSVs usam o prefixo ``_`` (``_motivo``,
    ``_status_http``...). O envio (SendApiTask) IGNORA colunas iniciadas
    por ``_`` ao montar o payload — entao o ``registros_falhos.csv`` pode
    ser usado DIRETO como nova planilha de entrada: o payload enviado e
    identico ao original, a Idempotency-Key e a MESMA, e nada duplica.

Nao inventa dado: apenas separa, anota e apresenta o que o envio produziu.
Estilo visual do XLSX segue o padrao do report_xlsx da Auditoria
(helpers duplicados de proposito para nao acoplar os dois fluxos).
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from autotarefas.core.base import TaskResult
from autotarefas.tasks.report import write_json_report

#: Nomes canonicos dos artefatos.
IMPORT_JSON_NAME = "importacao_report.json"
SENT_CSV_NAME = "registros_enviados.csv"
FAILED_CSV_NAME = "registros_falhos.csv"
RESULT_XLSX_NAME = "importacao_resultado.xlsx"

#: Rotulos legiveis das categorias de falha (Resumo/Falhas do XLSX).
_CATEGORY_LABELS = {
    "validacao": "Dados invalidos (400/422)",
    "duplicado": "Ja cadastrado (409)",
    "rate_limit": "Limite de requisicoes (429)",
    "temporario": "Instabilidade do sistema (5xx)",
    "conexao": "Falha de conexao",
    "outro": "Outros",
}

# ============================================================
# Estilo (mesmo vocabulario visual do report_xlsx da Auditoria)
# ============================================================

_FONT = "Arial"
_TITLE_FONT = Font(name=_FONT, bold=True, size=16, color="1F4E78")
_LABEL_FONT = Font(name=_FONT, bold=True)
_HEADER_FONT = Font(name=_FONT, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_OK_FILL = PatternFill("solid", fgColor="E2EFDA")
_ERROR_FILL = PatternFill("solid", fgColor="FCE4E4")
_NEUTRAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_CENTER = Alignment(horizontal="center", vertical="center")

_SHEET_RESUMO = "Resumo"
_SHEET_ENVIADOS = "Enviados"
_SHEET_FALHAS = "Falhas"
_SHEET_DETALHES = "Detalhes"


def _cell_value(value: object) -> object:
    """Converte celula do DataFrame para um tipo que o openpyxl aceita."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    if isinstance(value, (bool, int, float, str)):
        return value
    item = getattr(value, "item", None)
    if callable(item):
        native = item()
        if isinstance(native, (bool, int, float, str)):
            return native
    return str(value)


def _style_header_row(ws: Worksheet, n_cols: int) -> None:
    """Cabecalho formatado + painel congelado + autofiltro."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
    ws.freeze_panes = "A2"
    last_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def _autofit_columns(ws: Worksheet, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _write_dataframe(ws: Worksheet, dataframe: pd.DataFrame) -> None:
    ws.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        ws.append([_cell_value(v) for v in row])
    data_font = Font(name=_FONT)
    for cell_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in cell_row:
            cell.font = data_font
    _style_header_row(ws, len(dataframe.columns))
    _autofit_columns(ws)


# ============================================================
# Selecao de linhas e montagem dos CSVs
# ============================================================


def _drop_meta_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    Remove colunas de metadado (iniciadas por ``_``) da planilha de origem.

    Garante que reenviar um ``registros_falhos.csv`` gere artefatos
    limpos: os metadados sao sempre os da execucao ATUAL.
    """
    keep = [c for c in dataframe.columns if not str(c).startswith("_")]
    return dataframe[keep]


def _lines_to_indices(lines: list[int]) -> list[int]:
    """Linha fisica (cabecalho=1; 1a de dados=2) -> indice do DataFrame."""
    return [n - 2 for n in lines]


def _split_items(items: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Separa items em (enviados, falhos), preservando a ordem."""
    sent = [i for i in items if i.get("sucesso")]
    failed = [i for i in items if not i.get("sucesso")]
    return sent, failed


def write_sent_csv(dataframe: pd.DataFrame, items: list[dict[str, Any]], path: Path) -> None:
    """Escreve o CSV dos registros ENVIADOS (+ ``_status_http``/``_id_externo``)."""
    sent, _failed = _split_items(items)
    base = _drop_meta_columns(dataframe)
    subset = base.iloc[_lines_to_indices([int(i["linha"]) for i in sent])].copy()
    subset["_status_http"] = [i.get("status_http") for i in sent]
    subset["_id_externo"] = [i.get("id_externo") or "" for i in sent]
    path.parent.mkdir(parents=True, exist_ok=True)
    subset.to_csv(path, index=False, encoding="utf-8-sig")


def write_failed_csv(dataframe: pd.DataFrame, items: list[dict[str, Any]], path: Path) -> None:
    """
    Escreve o CSV dos registros FALHOS.

    Colunas originais + ``_status_http``, ``_categoria``, ``_motivo``,
    ``_tentativas`` e ``_pode_reenviar``. Como o envio ignora colunas
    ``_...``, este arquivo pode ser corrigido e usado DIRETO como nova
    entrada — com a mesma Idempotency-Key para o que nao mudou.
    """
    _sent, failed = _split_items(items)
    base = _drop_meta_columns(dataframe)
    subset = base.iloc[_lines_to_indices([int(i["linha"]) for i in failed])].copy()
    subset["_status_http"] = [i.get("status_http") for i in failed]
    subset["_categoria"] = [i.get("categoria") for i in failed]
    subset["_motivo"] = [i.get("mensagem") for i in failed]
    subset["_tentativas"] = [i.get("tentativas") for i in failed]
    subset["_pode_reenviar"] = [bool(i.get("pode_reenviar")) for i in failed]
    path.parent.mkdir(parents=True, exist_ok=True)
    subset.to_csv(path, index=False, encoding="utf-8-sig")


# ============================================================
# XLSX executivo (4 abas)
# ============================================================


def _build_resumo(ws: Worksheet, result: TaskResult) -> None:
    data = result.data
    ws["A1"] = "Cadastro automatico via planilha"
    ws["A1"].font = _TITLE_FONT

    rows: list[tuple[str, object]] = [
        ("Sistema (URL)", str(data.get("url", ""))),
        ("Planilha", str(data.get("planilha", ""))),
        ("Total de registros", int(data.get("total", 0))),
        ("Enviados", int(data.get("enviados", 0))),
        ("Falhos", int(data.get("falhas", 0))),
        ("Reenviaveis", int(data.get("reenviaveis", 0))),
    ]
    fills = {
        "Enviados": _OK_FILL,
        "Falhos": _ERROR_FILL,
        "Reenviaveis": _NEUTRAL_FILL,
    }
    start = 3
    for offset, (label, value) in enumerate(rows):
        r = start + offset
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = Font(name=_FONT)
        if label in fills:
            ws.cell(row=r, column=1).fill = fills[label]
            cell.fill = fills[label]

    header_row = start + len(rows) + 1
    ws.cell(row=header_row, column=1, value="Falhas por categoria").font = _LABEL_FONT
    ws.cell(row=header_row + 1, column=1, value="Categoria").font = _HEADER_FONT
    ws.cell(row=header_row + 1, column=1).fill = _HEADER_FILL
    ws.cell(row=header_row + 1, column=2, value="Quantidade").font = _HEADER_FONT
    ws.cell(row=header_row + 1, column=2).fill = _HEADER_FILL

    by_category: dict[str, int] = data.get("falhas_por_categoria", {})
    line = header_row + 2
    if by_category:
        for category, count in by_category.items():
            label = _CATEGORY_LABELS.get(category, category)
            ws.cell(row=line, column=1, value=label).font = Font(name=_FONT)
            ws.cell(row=line, column=2, value=count).font = Font(name=_FONT)
            line += 1
    else:
        ws.cell(row=line, column=1, value="Nenhuma falha").font = Font(name=_FONT)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70


def _build_detalhes(ws: Worksheet, items: list[dict[str, Any]]) -> None:
    """Log tecnico por item (linha, status, categoria, tentativas, chave)."""
    if not items:
        ws["A1"] = "Nenhum registro processado."
        ws["A1"].font = Font(name=_FONT)
        ws.column_dimensions["A"].width = 40
        return
    detalhes = pd.DataFrame(
        {
            "linha": [i.get("linha") for i in items],
            "status_http": [i.get("status_http") for i in items],
            "categoria": [i.get("categoria") for i in items],
            "tentativas": [i.get("tentativas") for i in items],
            "idempotency_key": [i.get("idempotency_key") for i in items],
            "mensagem": [i.get("mensagem") for i in items],
        }
    )
    _write_dataframe(ws, detalhes)


def write_result_xlsx(dataframe: pd.DataFrame, result: TaskResult, path: Path) -> None:
    """Gera o ``importacao_resultado.xlsx`` com as 4 abas."""
    items: list[dict[str, Any]] = result.data.get("items", [])
    sent, failed = _split_items(items)
    base = _drop_meta_columns(dataframe)

    wb = Workbook()
    resumo = wb.active
    resumo.title = _SHEET_RESUMO
    _build_resumo(resumo, result)

    enviados_df = base.iloc[_lines_to_indices([int(i["linha"]) for i in sent])].copy()
    enviados_df["_status_http"] = [i.get("status_http") for i in sent]
    enviados_df["_id_externo"] = [i.get("id_externo") or "" for i in sent]
    _write_dataframe(wb.create_sheet(_SHEET_ENVIADOS), enviados_df)

    falhas_df = base.iloc[_lines_to_indices([int(i["linha"]) for i in failed])].copy()
    falhas_df["_status_http"] = [i.get("status_http") for i in failed]
    falhas_df["_categoria"] = [
        _CATEGORY_LABELS.get(str(i.get("categoria")), str(i.get("categoria"))) for i in failed
    ]
    falhas_df["_motivo"] = [i.get("mensagem") for i in failed]
    falhas_df["_pode_reenviar"] = ["Sim" if i.get("pode_reenviar") else "Nao" for i in failed]
    _write_dataframe(wb.create_sheet(_SHEET_FALHAS), falhas_df)

    _build_detalhes(wb.create_sheet(_SHEET_DETALHES), items)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ============================================================
# Orquestrador
# ============================================================


def write_send_artifacts(
    dataframe: pd.DataFrame,
    result: TaskResult,
    out_dir: Path,
) -> tuple[Path, Path, Path, Path]:
    """
    Gera os 4 artefatos canonicos em ``out_dir``.

    Returns:
        Tupla ``(enviados_csv, falhos_csv, resultado_xlsx, report_json)``.
    """
    items: list[dict[str, Any]] = result.data.get("items", [])
    out_dir.mkdir(parents=True, exist_ok=True)

    sent_path = out_dir / SENT_CSV_NAME
    failed_path = out_dir / FAILED_CSV_NAME
    xlsx_path = out_dir / RESULT_XLSX_NAME
    json_path = out_dir / IMPORT_JSON_NAME

    write_sent_csv(dataframe, items, sent_path)
    write_failed_csv(dataframe, items, failed_path)
    write_result_xlsx(dataframe, result, xlsx_path)
    write_json_report(result, json_path)
    return sent_path, failed_path, xlsx_path, json_path


__all__ = [
    "FAILED_CSV_NAME",
    "IMPORT_JSON_NAME",
    "RESULT_XLSX_NAME",
    "SENT_CSV_NAME",
    "write_failed_csv",
    "write_result_xlsx",
    "write_send_artifacts",
    "write_sent_csv",
]
