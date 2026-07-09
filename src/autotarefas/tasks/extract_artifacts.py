"""
Artefatos da Exportacao automatica de dados (extract_api --out-dir).

A partir dos registros extraidos e do TaskResult, gera o pacote de
artefatos da Exportacao — o primeiro passo do pipeline do AutoTarefas
(Exportacao -> Auditoria -> Cadastro):

- ``dados_extraidos.csv``    a base extraida, pronta p/ importar/auditar
- ``dados_extraidos.xlsx``   a mesma base em Excel, para o cliente abrir
- ``extracao_report.json``   contadores da extracao (contrato do front)

Diferente da Auditoria e do Cadastro, a Exportacao NAO separa sucesso/
falha por linha — ela traz uma base. Por isso o pacote e mais enxuto
(3 artefatos) e o XLSX tem 2 abas (Resumo + Dados).

Estilo visual do XLSX segue o padrao do report_xlsx da Auditoria e do
send_artifacts do Cadastro (helpers duplicados de proposito para nao
acoplar os fluxos).
"""

from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from autotarefas.core.base import TaskResult

#: Nomes canonicos dos artefatos.
DATA_CSV_NAME = "dados_extraidos.csv"
DATA_XLSX_NAME = "dados_extraidos.xlsx"
EXTRACT_JSON_NAME = "extracao_report.json"

# ============================================================
# Estilo (mesmo vocabulario visual dos outros fluxos)
# ============================================================

_FONT = "Arial"
_TITLE_FONT = Font(name=_FONT, bold=True, size=16, color="1F4E78")
_LABEL_FONT = Font(name=_FONT, bold=True)
_HEADER_FONT = Font(name=_FONT, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_NEUTRAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_CENTER = Alignment(horizontal="center", vertical="center")

_SHEET_RESUMO = "Resumo"
_SHEET_DADOS = "Dados"


def _cell_value(value: object) -> object:
    """Converte celula do DataFrame para um tipo que o openpyxl aceita."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    if isinstance(value, (bool, int, float, str)):
        return value
    item = getattr(value, "item", None)
    if callable(item):
        native = item()
        if isinstance(native, (bool, int, float, str)):
            return native
    return str(value)


def _style_header_row(ws: Worksheet, n_cols: int) -> None:
    """Cabecalho formatado + painel congelado + autofiltro."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
    ws.freeze_panes = "A2"
    last_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def _autofit_columns(ws: Worksheet, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def _write_dataframe(ws: Worksheet, dataframe: pd.DataFrame) -> None:
    ws.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        ws.append([_cell_value(v) for v in row])
    data_font = Font(name=_FONT)
    for cell_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in cell_row:
            cell.font = data_font
    _style_header_row(ws, len(dataframe.columns))
    _autofit_columns(ws)


# ============================================================
# Report JSON (contrato do front)
# ============================================================


def build_report_data(records: list[dict[str, Any]], result: TaskResult) -> dict[str, Any]:
    """
    Monta o dicionario do extracao_report.json a partir do resultado.

    Campos consumidos pelo front (resumo visual): total_registros,
    paginas, colunas, formato_origem, origem.
    """
    data = result.data
    colunas = list(records[0].keys()) if records else []
    return {
        "total_registros": len(records),
        "paginas": data.get("total_pages"),
        "colunas": colunas,
        "origem": data.get("url", ""),
        "duracao_ms": result.duration_ms,
    }


# ============================================================
# XLSX (2 abas: Resumo + Dados)
# ============================================================


def _build_resumo(ws: Worksheet, report: dict[str, Any]) -> None:
    ws["A1"] = "Exportacao automatica de dados"
    ws["A1"].font = _TITLE_FONT

    paginas = report.get("paginas")
    rows: list[tuple[str, object]] = [
        ("Origem (URL)", str(report.get("origem", ""))),
        ("Registros extraidos", int(report.get("total_registros", 0))),
        ("Paginas percorridas", paginas if paginas is not None else "-"),
        ("Colunas", len(report.get("colunas", []))),
    ]
    start = 3
    for offset, (label, value) in enumerate(rows):
        r = start + offset
        ws.cell(row=r, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = Font(name=_FONT)
        if label == "Registros extraidos":
            ws.cell(row=r, column=1).fill = _NEUTRAL_FILL
            cell.fill = _NEUTRAL_FILL

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70


def write_data_xlsx(dataframe: pd.DataFrame, report: dict[str, Any], path: Path) -> None:
    """Gera o ``dados_extraidos.xlsx`` com as abas Resumo + Dados."""
    wb = Workbook()
    resumo = wb.active
    resumo.title = _SHEET_RESUMO
    _build_resumo(resumo, report)

    dados = wb.create_sheet(_SHEET_DADOS)
    if dataframe.empty:
        dados["A1"] = "Nenhum registro extraido."
        dados["A1"].font = Font(name=_FONT)
        dados.column_dimensions["A"].width = 40
    else:
        _write_dataframe(dados, dataframe)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ============================================================
# Orquestrador
# ============================================================


def write_extract_artifacts(
    records: list[dict[str, Any]],
    result: TaskResult,
    out_dir: Path,
) -> tuple[Path, Path, Path]:
    """
    Gera os 3 artefatos canonicos da Exportacao em ``out_dir``.

    Returns:
        Tupla ``(dados_csv, dados_xlsx, report_json)``.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    dataframe = pd.DataFrame(records)
    report = build_report_data(records, result)

    csv_path = out_dir / DATA_CSV_NAME
    xlsx_path = out_dir / DATA_XLSX_NAME
    json_path = out_dir / EXTRACT_JSON_NAME

    dataframe.to_csv(csv_path, index=False, encoding="utf-8-sig")
    write_data_xlsx(dataframe, report, xlsx_path)
    json_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return csv_path, xlsx_path, json_path


__all__ = [
    "DATA_CSV_NAME",
    "DATA_XLSX_NAME",
    "EXTRACT_JSON_NAME",
    "build_report_data",
    "write_data_xlsx",
    "write_extract_artifacts",
]
