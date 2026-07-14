"""
Leitura de planilhas: XLSX (openpyxl) e CSV.

Por que openpyxl e nao pandas, no XLSX: o pandas DESCARTA o
``number_format`` e o ``data_type`` da celula. Sem esses metadados, o
leitor teria que ADIVINHAR o que o arquivo ja AFIRMA — por exemplo, que
uma coluna e moeda (formato ``"R$" #,##0.00``). O pandas continua sendo
usado, mas DEPOIS: para montar o DataFrame a partir do que foi lido.

No CSV nao ha metadado nenhum (tudo e texto), entao la a inferencia e
obrigatoria.
"""

from __future__ import annotations

import csv
import zipfile
from pathlib import Path
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from autotarefas.reader.normalize import normalize_column
from autotarefas.reader.result import (
    ColumnInfo,
    Conversion,
    ReadWarning,
    SheetCandidate,
    WorkbookReadResult,
)
from autotarefas.reader.table_detect import (
    MIN_HEADER_CONFIDENCE,
    MIN_SHEET_SCORE,
    SHEET_MARGIN,
    Grid,
    find_header_row,
    looks_like_footer,
    score_sheet,
)
from autotarefas.reader.types import RawCell, infer_column_type

if TYPE_CHECKING:
    from autotarefas.reader.result import FileType

#: Limites operacionais (acima disso, recusa com mensagem clara).
MAX_ROWS = 200_000
MAX_BYTES = 50 * 1024 * 1024

_ENCODINGS = ("utf-8-sig", "utf-8", "latin-1")
_SAMPLE_SIZE = 5
_MAX_FORMULA_SCAN_ROWS = 50_000


class ReaderError(Exception):
    """Falha irrecuperavel ao abrir o arquivo."""


# --- Leitura bruta -----------------------------------------------------------


def _read_csv_grid(path: Path) -> Grid:
    """Le um CSV inteiro como texto (sem inferencia do pandas)."""
    texto = None
    for encoding in _ENCODINGS:
        try:
            texto = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        msg = "nao foi possivel decodificar o arquivo (tente salvar como UTF-8)"
        raise ReaderError(msg)

    amostra = texto[:8192]
    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=",;\t|")
        delimitador = dialect.delimiter
    except csv.Error:
        delimitador = ","

    linhas = list(csv.reader(texto.splitlines(), delimiter=delimitador))
    largura = max((len(r) for r in linhas), default=0)
    return [[RawCell(value=(r[i] if i < len(r) else None)) for i in range(largura)] for r in linhas]


def _read_xlsx_grids(path: Path) -> tuple[dict[str, Grid], set[tuple[str, int, int]]]:
    """
    Le todas as abas de um XLSX preservando tipo e formato de cada celula.

    Returns:
        (grades por aba, celulas que eram FORMULA).
    """
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError) as exc:
        msg = f"nao foi possivel abrir o arquivo: {exc}"
        raise ReaderError(msg) from exc

    grids: dict[str, Grid] = {}
    try:
        for ws in wb.worksheets:
            grid: Grid = []
            for row in ws.iter_rows():
                grid.append(
                    [
                        RawCell(
                            value=cell.value,
                            excel_type=str(cell.data_type or ""),
                            number_format=str(cell.number_format or ""),
                        )
                        for cell in row
                    ]
                )
            grids[str(ws.title)] = grid
    finally:
        wb.close()

    return grids, _scan_formulas(path, grids)


def _tem_formulas(path: Path) -> bool:
    """
    O arquivo contem alguma formula? Checagem BARATA, antes de reler tudo.

    O .xlsx e um zip: a formula aparece como a tag <f> dentro do XML da aba.
    Uma busca de bytes custa milissegundos — abrir o arquivo inteiro de novo
    com o openpyxl custa segundos. A maioria das planilhas nao tem formula
    nenhuma, e nao deve pagar por isso.
    """
    try:
        with zipfile.ZipFile(path) as z:
            nomes = [n for n in z.namelist() if n.startswith("xl/worksheets/sheet")]
            return any(b"<f" in z.read(nome) for nome in nomes)
    except (zipfile.BadZipFile, KeyError, OSError):
        return False


def _scan_formulas(path: Path, grids: dict[str, Grid]) -> set[tuple[str, int, int]]:
    """Segunda passada: quais celulas eram formula (data_only=False)."""
    total = sum(len(g) for g in grids.values())
    if total > _MAX_FORMULA_SCAN_ROWS or not _tem_formulas(path):
        return set()

    formulas: set[tuple[str, int, int]] = set()
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except (InvalidFileException, KeyError, OSError):
        return set()
    try:
        for ws in wb.worksheets:
            for r, row in enumerate(ws.iter_rows()):
                for c, cell in enumerate(row):
                    if cell.data_type == "f":
                        formulas.add((str(ws.title), r, c))
    finally:
        wb.close()
    return formulas


# --- Selecao de aba ----------------------------------------------------------


def _rank_sheets(grids: dict[str, Grid]) -> list[SheetCandidate]:
    candidatos: list[SheetCandidate] = []
    for nome, grid in grids.items():
        score, motivo = score_sheet(grid)
        larguras = [sum(1 for c in r if not c.is_empty) for r in grid]
        candidatos.append(
            SheetCandidate(
                name=nome,
                score=score,
                reason=motivo,
                rows=sum(1 for w in larguras if w > 0),
                cols=max(larguras, default=0),
            )
        )
    candidatos.sort(key=lambda s: -s.score)
    return candidatos


def _select_sheet(
    candidatos: list[SheetCandidate],
    solicitada: str | None,
) -> tuple[str | None, float, str | None]:
    """Retorna (aba, confianca, motivo_de_recusa)."""
    if solicitada is not None:
        for cand in candidatos:
            if cand.name == solicitada:
                return cand.name, 1.0, None
        nomes = ", ".join(c.name for c in candidatos)
        return None, 0.0, f"aba '{solicitada}' nao existe. Abas disponiveis: {nomes}"

    if not candidatos:
        return None, 0.0, "o arquivo nao tem nenhuma aba com dados"

    melhor = candidatos[0]
    if melhor.score < MIN_SHEET_SCORE:
        return (
            None,
            melhor.score,
            (
                "nenhuma aba parece uma tabela (com linha de cabecalho e registros em linhas). "
                f"Melhor candidata: '{melhor.name}' ({melhor.reason})"
            ),
        )

    empatadas = [c for c in candidatos[1:] if melhor.score - c.score < SHEET_MARGIN]
    if empatadas:
        opcoes = ", ".join(f"'{c.name}' ({c.score:.2f})" for c in [melhor, *empatadas])
        return (
            None,
            melhor.score,
            (f"mais de uma aba parece uma tabela: {opcoes}. Use --sheet para escolher."),
        )

    return melhor.name, melhor.score, None


# --- Montagem do resultado ---------------------------------------------------


def _unique_headers(linha: list[RawCell]) -> list[str]:
    """Nomes das colunas, preservando o original e desambiguando repetidos."""
    nomes: list[str] = []
    vistos: dict[str, int] = {}
    for i, cell in enumerate(linha):
        base = cell.text or f"coluna_{i + 1}"
        if base in vistos:
            vistos[base] += 1
            base = f"{base}_{vistos[base]}"
        else:
            vistos[base] = 0
        nomes.append(base)
    return nomes


def _rejected(path: Path, file_type: FileType, motivo: str, **extra: object) -> WorkbookReadResult:
    return WorkbookReadResult(
        source_file=path,
        file_type=file_type,
        rejected_reason=motivo,
        **extra,  # type: ignore[arg-type]
    )


def read_workbook(
    path: Path,
    sheet: str | None = None,
    header_row: int | None = None,
) -> WorkbookReadResult:
    """
    Le uma planilha (CSV ou XLSX) e devolve o resultado completo.

    Args:
        path: Arquivo a ler.
        sheet: Aba a usar (XLSX). None = detectar.
        header_row: Linha FISICA do cabecalho (1-based). None = detectar.

    Returns:
        WorkbookReadResult. Se `rejected_reason` estiver preenchido, o
        arquivo NAO foi processado — e o motivo esta la, em portugues.
    """
    if not path.exists():
        msg = f"arquivo nao encontrado: {path}"
        raise ReaderError(msg)
    if path.stat().st_size > MAX_BYTES:
        return _rejected(path, "xlsx", f"arquivo maior que {MAX_BYTES // 1024 // 1024} MB")

    sufixo = path.suffix.lower()
    if sufixo == ".csv":
        file_type: FileType = "csv"
        grids = {"csv": _read_csv_grid(path)}
        formulas: set[tuple[str, int, int]] = set()
    elif sufixo in (".xlsx", ".xlsm"):
        file_type = "xlsx"
        grids, formulas = _read_xlsx_grids(path)
    elif sufixo == ".xls":
        return _rejected(path, "xlsx", "formato .xls antigo nao suportado: converta para .xlsx")
    else:
        return _rejected(path, "csv", f"extensao nao suportada: {sufixo} (use .csv ou .xlsx)")

    candidatos = _rank_sheets(grids)
    nome_aba, conf_aba, recusa = _select_sheet(candidatos, sheet)
    if nome_aba is None:
        return _rejected(path, file_type, recusa or "aba indefinida", available_sheets=candidatos)

    grid = grids[nome_aba]
    return _build_result(
        path=path,
        file_type=file_type,
        grid=grid,
        sheet_name=nome_aba,
        sheet_conf=conf_aba,
        candidatos=candidatos,
        header_row=header_row,
        formulas={(r, c) for (s, r, c) in formulas if s == nome_aba},
    )


def _build_result(  # noqa: PLR0913
    *,
    path: Path,
    file_type: FileType,
    grid: Grid,
    sheet_name: str,
    sheet_conf: float,
    candidatos: list[SheetCandidate],
    header_row: int | None,
    formulas: set[tuple[int, int]],
) -> WorkbookReadResult:
    """Detecta cabecalho, tipa as colunas, normaliza e monta o resultado."""
    idx_header: int | None
    if header_row is not None:
        idx_header = header_row - 1
        conf_header = 1.0
        alternativas: list[int] = []
        if idx_header < 0 or idx_header >= len(grid):
            return _rejected(
                path,
                file_type,
                f"--header-row {header_row} fora do arquivo (ele tem {len(grid)} linhas)",
                available_sheets=candidatos,
                selected_sheet=sheet_name,
            )
    else:
        idx_header, conf_header, alternativas = find_header_row(grid)
        if idx_header is None or conf_header < MIN_HEADER_CONFIDENCE:
            return _rejected(
                path,
                file_type,
                (
                    "nao encontrei uma linha de cabecalho confiavel: o arquivo nao parece "
                    "tabular. Use --header-row para indicar a linha do cabecalho."
                ),
                available_sheets=candidatos,
                selected_sheet=sheet_name,
                header_confidence=conf_header,
            )

    linha_header = grid[idx_header]
    nomes = _unique_headers(linha_header)
    n_cols = len(nomes)

    regiao = grid[idx_header + 1 :]
    corpo = [r for r in regiao if any(not c.is_empty for c in r)]
    vazias_ignoradas = len(regiao) - len(corpo)
    if not corpo:
        return _rejected(
            path,
            file_type,
            "o cabecalho foi encontrado, mas nao ha nenhuma linha de dados abaixo dele",
            available_sheets=candidatos,
            selected_sheet=sheet_name,
            header_row=idx_header + 1,
        )

    avisos: list[ReadWarning] = []
    linha_dados_inicio = idx_header + 2  # fisica, 1-based

    # rodape de totais: MARCA, nunca remove
    if looks_like_footer(grid, len(grid) - 1, n_cols):
        avisos.append(
            ReadWarning(
                code="rodape_suspeito",
                message=(
                    "a ultima linha parece um rodape de totais (nao foi removida; "
                    "confira antes de analisar)"
                ),
                row=len(grid),
            )
        )

    if vazias_ignoradas:
        avisos.append(
            ReadWarning(
                code="linhas_vazias_ignoradas",
                message=(
                    f"{vazias_ignoradas} linha(s) totalmente vazia(s) nao foram contadas "
                    "como registro"
                ),
            )
        )

    if formulas:
        avisos.append(
            ReadWarning(
                code="formulas",
                message=(
                    f"{len(formulas)} celula(s) continham formula; foi usado o valor "
                    "calculado que estava salvo no arquivo"
                ),
            )
        )

    colunas, original, normalizado, conversoes, avisos_col = _process_columns(
        corpo, nomes, linha_dados_inicio
    )
    avisos.extend(avisos_col)
    avisos.extend(_structural_warnings(original))

    confianca = round(min(sheet_conf, conf_header), 3)
    return WorkbookReadResult(
        source_file=path,
        file_type=file_type,
        available_sheets=candidatos,
        selected_sheet=sheet_name,
        sheet_confidence=sheet_conf,
        header_row=idx_header + 1,
        header_confidence=conf_header,
        header_alternatives=[i + 1 for i in alternativas],
        data_start_row=linha_dados_inicio,
        data_end_row=linha_dados_inicio + len(corpo) - 1,
        detected_columns=colunas,
        original_dataframe=original,
        normalized_dataframe=normalizado,
        conversions=conversoes,
        warnings=avisos,
        skipped_empty_rows=vazias_ignoradas,
        confidence=confianca,
    )


def _process_columns(
    corpo: Grid,
    nomes: list[str],
    primeira_linha: int,
) -> tuple[list[ColumnInfo], pd.DataFrame, pd.DataFrame, list[Conversion], list[ReadWarning]]:
    """Tipa, normaliza e monta os dois DataFrames (fiel e tipado)."""
    colunas: list[ColumnInfo] = []
    dados_originais: dict[str, list[str]] = {}
    dados_normalizados: dict[str, list[object]] = {}
    conversoes: list[Conversion] = []
    avisos: list[ReadWarning] = []

    for i, nome in enumerate(nomes):
        celulas = [(r[i] if i < len(r) else RawCell(value=None)) for r in corpo]
        typing = infer_column_type(celulas)
        tipo, conf, observacoes = typing.inferred_type, typing.confidence, typing.observations

        formatos = {c.number_format for c in celulas if c.number_format}
        formato = next(iter(formatos), None) if len(formatos) == 1 else None

        colunas.append(
            ColumnInfo(
                index=i,
                name=nome,
                inferred_type=tipo,
                type_confidence=round(conf, 3),
                observations=observacoes,
                excel_number_format=formato,
                sample_values=[c.text for c in celulas if not c.is_empty][:_SAMPLE_SIZE],
                empty_count=sum(1 for c in celulas if c.is_empty),
                type_counts=typing.type_counts,
            )
        )

        if tipo == "misto":
            avisos.append(
                ReadWarning(
                    code="coluna_mista",
                    message=(
                        f"coluna '{nome}': {observacoes[0] if observacoes else 'tipos misturados'}"
                    ),
                    column=nome,
                )
            )
        if tipo == "erro":
            avisos.append(
                ReadWarning(
                    code="erro_excel",
                    message=f"coluna '{nome}' contem valores de erro do Excel (#DIV/0!, #REF!...)",
                    column=nome,
                )
            )

        # o original guarda o texto EXATO do arquivo, COM os espacos
        dados_originais[nome] = [c.raw for c in celulas]
        valores, convs = normalize_column(celulas, tipo, nome, primeira_linha)
        dados_normalizados[nome] = valores
        conversoes.extend(convs)

    return (
        colunas,
        pd.DataFrame(dados_originais, dtype=str),
        pd.DataFrame(dados_normalizados),
        conversoes,
        avisos,
    )


def _structural_warnings(original: pd.DataFrame) -> list[ReadWarning]:
    """Avisos que valem para a tabela toda (nada e removido)."""
    avisos: list[ReadWarning] = []

    duplicadas = int(original.duplicated().sum())
    if duplicadas:
        avisos.append(
            ReadWarning(
                code="linhas_duplicadas",
                message=(
                    f"{duplicadas} linha(s) completamente identica(s) a outra (nao foram removidas)"
                ),
            )
        )

    vazias = [str(c) for c in original.columns if (original[c] == "").all()]
    for nome in vazias:
        avisos.append(
            ReadWarning(code="coluna_vazia", message=f"coluna '{nome}' esta vazia", column=nome)
        )

    return avisos


__all__ = ["MAX_BYTES", "MAX_ROWS", "ReaderError", "read_workbook"]
