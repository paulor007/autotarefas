"""
Gerador das fixtures de planilha.

Cada fixture isola UM problema. Elas sao pequenas (<5 KB), deterministicas
e versionadas — sao a defesa contra o leitor ficar "bom na planilha do
Paulo e ruim no resto do mundo".

Rodar:  python tests/fixtures/planilhas/build_fixtures.py
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

AQUI = Path(__file__).parent


def _xlsx(nome: str, linhas: list[list[object]], aba: str = "Plan1") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = aba
    for linha in linhas:
        ws.append(linha)
    wb.save(AQUI / nome)


def _csv(nome: str, linhas: list[list[str]], delim: str = ",", encoding: str = "utf-8") -> None:
    with (AQUI / nome).open("w", newline="", encoding=encoding) as f:
        csv.writer(f, delimiter=delim).writerows(linhas)


# --- LIMPAS ------------------------------------------------------------------


def f01_csv_limpo() -> None:
    _csv(
        "01_csv_limpo.csv",
        [
            ["nome", "idade", "cidade"],
            ["Ana", "34", "Sao Paulo"],
            ["Bruno", "28", "Recife"],
            ["Carla", "41", "Curitiba"],
        ],
    )


def f02_xlsx_limpo() -> None:
    _xlsx(
        "02_xlsx_limpo.xlsx",
        [
            ["nome", "idade", "cidade"],
            ["Ana", 34, "Sao Paulo"],
            ["Bruno", 28, "Recife"],
            ["Carla", 41, "Curitiba"],
        ],
    )


def f03_clientes() -> None:
    _xlsx(
        "03_clientes.xlsx",
        [
            ["nome", "email", "telefone", "cpf"],
            ["Ana Lima", "ana@x.com", "(11) 98888-0001", "529.982.247-25"],
            ["Bruno Sa", "bruno@x.com", "(21) 3344-5566", "168.995.350-09"],
        ],
    )


def f04_vendas() -> None:
    _xlsx(
        "04_vendas.xlsx",
        [
            ["Codigo Venda", "Data", "Loja", "Produto", "Quantidade", "Valor Unitario"],
            [65014, datetime(2019, 12, 1), "Morumbi", "Sunga", 5, 114.90],
            [65014, datetime(2019, 12, 1), "Morumbi", "Casaco", 1, 269.50],
            [65016, datetime(2019, 12, 2), "Iguatemi", "Sapato", 2, 363.25],
        ],
    )


def f05_produtos() -> None:
    _xlsx(
        "05_produtos.xlsx",
        [
            ["sku", "produto", "preco", "estoque"],
            ["PRD-0001", "Camisa", 89.9, 12],
            ["PRD-0002", "Calca", 149.9, 0],
        ],
    )


# --- ESTRUTURA ---------------------------------------------------------------


def f06_cabecalho_linha_4() -> None:
    _xlsx(
        "06_cabecalho_linha_4.xlsx",
        [
            ["Relatorio de Vendas"],
            ["Gerado em 01/12/2019"],
            [],
            ["produto", "quantidade", "valor"],
            ["Camisa", 2, 100.0],
            ["Calca", 1, 150.0],
            ["Meia", 3, 20.0],
        ],
    )


def f07_tres_abas() -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Vendas"
    for linha in [
        ["produto", "qtd", "valor"],
        ["Camisa", 2, 100],
        ["Calca", 1, 150],
        ["Meia", 4, 20],
    ]:
        ws1.append(linha)
    ws2 = wb.create_sheet("Estoque")
    for linha in [
        ["sku", "saldo"],
        ["A1", 10],
        ["A2", 5],
        ["A3", 8],
    ]:
        ws2.append(linha)
    ws3 = wb.create_sheet("Clientes")
    for linha in [
        ["nome", "email"],
        ["Ana", "ana@x.com"],
        ["Bruno", "bruno@x.com"],
        ["Carla", "carla@x.com"],
    ]:
        ws3.append(linha)
    wb.save(AQUI / "07_tres_abas.xlsx")


def f08_capa_e_dados() -> None:
    wb = Workbook()
    capa = wb.active
    capa.title = "Capa"
    capa["A1"] = "RELATORIO MENSAL"
    capa["A3"] = "Empresa Exemplo Ltda"
    capa["A5"] = "Dezembro/2019"

    dados = wb.create_sheet("Dados")
    for linha in [
        ["produto", "quantidade", "valor"],
        ["Camisa", 2, 100.0],
        ["Calca", 1, 150.0],
        ["Meia", 3, 20.0],
        ["Tenis", 1, 299.0],
    ]:
        dados.append(linha)
    wb.save(AQUI / "08_capa_e_dados.xlsx")


def f24_mescladas_antes_do_cabecalho() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan1"
    ws["A1"] = "RELATORIO CONSOLIDADO"
    ws.merge_cells("A1:D1")
    ws.append([])
    for linha in [
        ["produto", "quantidade", "valor", "loja"],
        ["Camisa", 2, 100.0, "Centro"],
        ["Calca", 1, 150.0, "Norte"],
        ["Meia", 3, 20.0, "Sul"],
    ]:
        ws.append(linha)
    wb.save(AQUI / "24_mescladas_antes_do_cabecalho.xlsx")


def f25_nao_tabular() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Painel"
    ws["A1"] = "PAINEL EXECUTIVO"
    ws["A3"] = "Faturamento"
    ws["C3"] = 1_240_000
    ws["A5"] = "Margem"
    ws["C5"] = "18%"
    ws["F2"] = "Observacoes"
    ws["F4"] = "Fechamento parcial"
    wb.save(AQUI / "25_nao_tabular.xlsx")


# --- COLUNAS -----------------------------------------------------------------


def f09_ordem_diferente() -> None:
    _xlsx(
        "09_ordem_diferente.xlsx",
        [
            ["cpf", "nome", "telefone", "email"],
            ["529.982.247-25", "Ana", "(11) 98888-0001", "ana@x.com"],
            ["168.995.350-09", "Bruno", "(21) 3344-5566", "bruno@x.com"],
        ],
    )


def f10_colunas_desconhecidas() -> None:
    _xlsx(
        "10_colunas_desconhecidas.xlsx",
        [
            ["ref_interna", "xpto", "flag_z", "obs"],
            ["A-1", "azul", "S", "ok"],
            ["A-2", "verde", "N", "revisar"],
        ],
    )


def f11_coluna_vazia() -> None:
    _xlsx(
        "11_coluna_vazia.xlsx",
        [
            ["produto", "reservado", "valor"],
            ["Camisa", None, 100.0],
            ["Calca", None, 150.0],
            ["Meia", None, 20.0],
        ],
    )


# --- LINHAS ------------------------------------------------------------------


def f12_linhas_vazias_no_meio() -> None:
    _xlsx(
        "12_linhas_vazias_no_meio.xlsx",
        [
            ["produto", "quantidade", "valor"],
            ["Camisa", 2, 100.0],
            [],
            ["Calca", 1, 150.0],
            [],
            ["Meia", 3, 20.0],
        ],
    )


def f21_linhas_duplicadas() -> None:
    _xlsx(
        "21_linhas_duplicadas.xlsx",
        [
            ["produto", "quantidade", "valor"],
            ["Camisa", 2, 100.0],
            ["Calca", 1, 150.0],
            ["Camisa", 2, 100.0],
            ["Meia", 3, 20.0],
        ],
    )


def f23_rodape_de_totais() -> None:
    _xlsx(
        "23_rodape_de_totais.xlsx",
        [
            ["produto", "quantidade", "valor"],
            ["Camisa", 2, 100.0],
            ["Calca", 1, 150.0],
            ["Meia", 3, 20.0],
            ["TOTAL", 6, 270.0],
        ],
    )


# --- TIPOS -------------------------------------------------------------------


def f13_numero_como_texto() -> None:
    _csv(
        "13_numero_como_texto.csv",
        [
            ["produto", "quantidade", "valor"],
            ["Camisa", "2", "100"],
            ["Calca", "1", "150"],
            ["Meia", "3", "20"],
        ],
    )


def f14_moeda_br() -> None:
    _csv(
        "14_moeda_br.csv",
        [
            ["produto", "preco"],
            ["Camisa", "R$ 1.234,56"],
            ["Calca", "R$ 89,90"],
            ["Meia", "R$ 12,00"],
        ],
    )


def f15_moeda_us() -> None:
    _csv(
        "15_moeda_us.csv",
        [
            ["produto", "preco"],
            ["Camisa", "$1,234.56"],
            ["Calca", "$89.90"],
            ["Meia", "$12.00"],
        ],
    )


def f16_datas_seriais() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan1"
    ws.append(["produto", "data"])
    for produto, serial in [("Camisa", 43800), ("Calca", 43801), ("Meia", 43802)]:
        ws.append([produto, serial])  # numero cru, sem formato de data
    wb.save(AQUI / "16_datas_seriais.xlsx")


def f17_datas_ddmmaaaa() -> None:
    _csv(
        "17_datas_ddmmaaaa.csv",
        [
            ["produto", "data"],
            ["Camisa", "01/12/2019"],
            ["Calca", "15/12/2019"],
            ["Meia", "31/12/2019"],
        ],
    )


def f18_tipos_misturados() -> None:
    _csv(
        "18_tipos_misturados.csv",
        [
            ["produto", "quantidade"],
            ["Camisa", "2"],
            ["Calca", "uma duzia"],
            ["Meia", "3"],
        ],
    )


def f26_zeros_a_esquerda() -> None:
    _csv(
        "26_zeros_a_esquerda.csv",
        [
            ["codigo", "produto"],
            ["00123", "Camisa"],
            ["00456", "Calca"],
            ["00789", "Meia"],
        ],
    )


# --- FORMULAS E ERROS --------------------------------------------------------


def f19_formulas() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan1"
    ws.append(["produto", "quantidade", "valor_unitario", "total"])
    for i, (p, q, v) in enumerate([("Camisa", 2, 100), ("Calca", 1, 150), ("Meia", 3, 20)], 2):
        ws.append([p, q, v, f"=B{i}*C{i}"])
    wb.save(AQUI / "19_formulas.xlsx")


def f20_erros_excel() -> None:
    _xlsx(
        "20_erros_excel.xlsx",
        [
            ["produto", "margem"],
            ["Camisa", "#DIV/0!"],
            ["Calca", "#REF!"],
            ["Meia", "#N/D"],
        ],
    )


# --- DOMINIO -----------------------------------------------------------------


def f22_codigo_repetido_valido() -> None:
    _xlsx(
        "22_codigo_repetido_valido.xlsx",
        [
            ["codigo_venda", "produto", "quantidade"],
            [1001, "Camisa", 2],
            [1001, "Calca", 1],
            [1001, "Meia", 3],
            [1002, "Tenis", 1],
        ],
    )


def f27_valor_derivado_divergente() -> None:
    _xlsx(
        "27_valor_derivado_divergente.xlsx",
        [
            ["produto", "quantidade", "valor_unitario", "valor_final"],
            ["Camisa", 2, 100.0, 200.0],
            ["Calca", 3, 150.0, 400.0],  # <- deveria ser 450: DIVERGENTE
            ["Meia", 1, 20.0, 20.0],
        ],
    )


# --- VOLUME (espelha a estrutura de uma planilha empresarial real) -----------


def f28_vendas_sintetica() -> None:
    """
    Planilha de vendas SINTETICA, com volume e as mesmas classes de anomalia
    de uma exportacao empresarial real.

    IMPORTANTE: nao e uma amostra de arquivo de cliente. Os dados sao
    inventados por aritmetica deterministica (nenhum dado identificavel,
    nenhuma dependencia de random). Ela existe para provar que o leitor
    aguenta volume e a estrutura tipica — nunca para o leitor "aprender"
    o dominio de vendas.

    Anomalias embutidas (para as subetapas 1.5/1.6 exercitarem depois):
      - vendas multi-item (mesmo codigo em varias linhas)
      - 3 linhas completamente duplicadas
      - 2 codigos com Data divergente dentro do mesmo grupo
    """
    lojas = ["Loja Alfa", "Loja Beta", "Loja Gama", "Loja Delta", "Loja Epsilon"]
    produtos = ["Camisa", "Calca", "Meia", "Tenis", "Casaco", "Sunga", "Boné", "Cinto"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan1"
    ws.append(
        [
            "Codigo Venda",
            "Data",
            "ID Loja",
            "Produto",
            "Quantidade",
            "Valor Unitario",
            "Valor Final",
        ]
    )

    linhas: list[list[object]] = []
    codigo = 70001
    for i in range(120):  # 120 vendas
        dia = (i % 26) + 1
        loja = lojas[i % len(lojas)]
        itens = (i % 3) + 1  # 1 a 3 itens por venda
        for j in range(itens):
            produto = produtos[(i + j) % len(produtos)]
            qtd = ((i + j) % 5) + 1
            unitario = float(30 + ((i * 7 + j * 13) % 72) * 10)
            linhas.append(
                [
                    codigo,
                    datetime(2019, 12, dia),
                    loja,
                    produto,
                    qtd,
                    unitario,
                    round(qtd * unitario, 2),
                ]
            )
        codigo += 1

    # anomalia 1: 3 linhas completamente duplicadas
    linhas.extend([list(linhas[10]), list(linhas[42]), list(linhas[77])])

    # anomalia 2: 2 codigos com Data divergente DENTRO do grupo
    linhas.append([linhas[0][0], datetime(2019, 12, 28), linhas[0][2], "Meia", 1, 50.0, 50.0])
    linhas.append([linhas[5][0], datetime(2019, 12, 29), linhas[5][2], "Cinto", 2, 40.0, 80.0])

    moeda = '"R$"\\ #,##0.00'
    for linha in linhas:
        ws.append(linha)
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = moeda

    wb.save(AQUI / "28_vendas_sintetica.xlsx")


FIXTURES = [
    f01_csv_limpo,
    f02_xlsx_limpo,
    f03_clientes,
    f04_vendas,
    f05_produtos,
    f06_cabecalho_linha_4,
    f07_tres_abas,
    f08_capa_e_dados,
    f09_ordem_diferente,
    f10_colunas_desconhecidas,
    f11_coluna_vazia,
    f12_linhas_vazias_no_meio,
    f13_numero_como_texto,
    f14_moeda_br,
    f15_moeda_us,
    f16_datas_seriais,
    f17_datas_ddmmaaaa,
    f18_tipos_misturados,
    f19_formulas,
    f20_erros_excel,
    f21_linhas_duplicadas,
    f22_codigo_repetido_valido,
    f23_rodape_de_totais,
    f24_mescladas_antes_do_cabecalho,
    f25_nao_tabular,
    f26_zeros_a_esquerda,
    f27_valor_derivado_divergente,
    f28_vendas_sintetica,
]


def main() -> None:
    for func in FIXTURES:
        func()
    print(f"{len(FIXTURES)} fixtures geradas em {AQUI}")


if __name__ == "__main__":
    main()
