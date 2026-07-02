"""Testes para autotarefas.tasks.report_xlsx."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from autotarefas.core.base import TaskResult
from autotarefas.tasks.report_xlsx import XLSX_NAME, write_xlsx_report
from autotarefas.tasks.validate import ColumnSchema, Schema, ValidateTask

_CSV = "nome,email,cpf\nAna Lima,Ana@Example.COM,52998224725\nBruno,bruno-invalido,111.111.111-11\n"


def _run(tmp_path: Path) -> tuple[pd.DataFrame, TaskResult]:
    csv = tmp_path / "dados.csv"
    csv.write_text(_CSV, encoding="utf-8")
    schema = Schema(
        columns=[
            ColumnSchema(name="nome", min_length=3),
            ColumnSchema(name="email", format="email"),
            ColumnSchema(name="cpf", validator_br="cpf"),
        ]
    )
    task = ValidateTask(file_path=csv, schema=schema, mode="limpeza")
    result = task.run()
    assert task.processed_dataframe is not None
    return task.processed_dataframe, result


class TestWriteXlsxReport:
    def test_gera_quatro_abas(self, tmp_path: Path) -> None:
        df, result = _run(tmp_path)
        out = tmp_path / XLSX_NAME
        write_xlsx_report(df, result, out)
        assert out.exists()
        wb = load_workbook(out)
        assert wb.sheetnames == [
            "Resumo",
            "Registros validos",
            "Registros invalidos",
            "Auditoria",
        ]

    def test_resumo_tem_contadores(self, tmp_path: Path) -> None:
        df, result = _run(tmp_path)
        out = tmp_path / XLSX_NAME
        write_xlsx_report(df, result, out)
        ws = load_workbook(out)["Resumo"]
        labels = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(1, 20)
        }
        assert labels.get("Total de registros") == 2
        assert labels.get("Registros validos") == 1
        assert labels.get("Registros invalidos") == 1

    def test_invalidos_tem_coluna_motivo(self, tmp_path: Path) -> None:
        df, result = _run(tmp_path)
        out = tmp_path / XLSX_NAME
        write_xlsx_report(df, result, out)
        ws = load_workbook(out)["Registros invalidos"]
        assert "motivo" in [c.value for c in ws[1]]
        assert ws.max_row == 2  # cabecalho + 1 invalido (Bruno)

    def test_validos_saem_normalizados(self, tmp_path: Path) -> None:
        df, result = _run(tmp_path)
        out = tmp_path / XLSX_NAME
        write_xlsx_report(df, result, out)
        ws = load_workbook(out)["Registros validos"]
        row = [c.value for c in ws[2]]
        assert "ana@example.com" in row
        assert "529.982.247-25" in row

    def test_auditoria_tem_cabecalho_e_registros(self, tmp_path: Path) -> None:
        df, result = _run(tmp_path)
        out = tmp_path / XLSX_NAME
        write_xlsx_report(df, result, out)
        ws = load_workbook(out)["Auditoria"]
        assert [c.value for c in ws[1]] == ["linha", "coluna", "antes", "depois", "regras"]
        assert ws.max_row >= 2  # houve normalizacoes
