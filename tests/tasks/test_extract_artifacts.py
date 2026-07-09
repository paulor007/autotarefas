"""Testes para autotarefas.tasks.extract_artifacts (os 3 artefatos)."""

from __future__ import annotations

import json
from typing import TYPE_CHECKING, Any

import pandas as pd
from openpyxl import load_workbook

from autotarefas.core.base import TaskResult, TaskStatus
from autotarefas.tasks.extract_artifacts import (
    DATA_CSV_NAME,
    DATA_XLSX_NAME,
    EXTRACT_JSON_NAME,
    build_report_data,
    write_extract_artifacts,
)

if TYPE_CHECKING:
    from pathlib import Path

URL = "http://test.local/api/clientes"


def _records(n: int) -> list[dict[str, Any]]:
    return [{"id": i, "nome": f"Cliente {i}", "email": f"c{i}@x.com"} for i in range(1, n + 1)]


def _result(records: list[dict[str, Any]], total_pages: int = 2) -> TaskResult:
    from datetime import UTC, datetime

    now = datetime.now(UTC)
    return TaskResult(
        task_name="extract_api",
        status=TaskStatus.SUCCESS,
        started_at=now,
        finished_at=now,
        rows_affected=len(records),
        duration_ms=123,
        data={
            "extracted": len(records),
            "url": URL,
            "total_pages": total_pages,
            "saved": True,
        },
    )


class TestReportData:
    def test_campos_do_report(self) -> None:
        records = _records(15)
        report = build_report_data(records, _result(records, total_pages=2))
        assert report["total_registros"] == 15
        assert report["paginas"] == 2
        assert report["colunas"] == ["id", "nome", "email"]
        assert report["origem"] == URL

    def test_report_vazio(self) -> None:
        report = build_report_data([], _result([], total_pages=1))
        assert report["total_registros"] == 0
        assert report["colunas"] == []


class TestArtefatos:
    def test_gera_os_tres_arquivos(self, tmp_path: Path) -> None:
        records = _records(12)
        paths = write_extract_artifacts(records, _result(records), tmp_path / "saida")
        for p in paths:
            assert p.exists()
        nomes = sorted(p.name for p in paths)
        assert nomes == sorted([DATA_CSV_NAME, DATA_XLSX_NAME, EXTRACT_JSON_NAME])

    def test_csv_tem_todos_os_registros(self, tmp_path: Path) -> None:
        records = _records(23)
        write_extract_artifacts(records, _result(records), tmp_path / "saida")
        df = pd.read_csv(tmp_path / "saida" / DATA_CSV_NAME)
        assert len(df) == 23
        assert list(df.columns) == ["id", "nome", "email"]

    def test_json_report_valido(self, tmp_path: Path) -> None:
        records = _records(5)
        write_extract_artifacts(records, _result(records, total_pages=1), tmp_path / "saida")
        report = json.loads((tmp_path / "saida" / EXTRACT_JSON_NAME).read_text(encoding="utf-8"))
        assert report["total_registros"] == 5
        assert report["paginas"] == 1

    def test_xlsx_duas_abas_e_resumo(self, tmp_path: Path) -> None:
        records = _records(8)
        write_extract_artifacts(records, _result(records, total_pages=2), tmp_path / "saida")
        wb = load_workbook(tmp_path / "saida" / DATA_XLSX_NAME)
        assert wb.sheetnames == ["Resumo", "Dados"]

        resumo = wb["Resumo"]
        labels = {
            resumo.cell(row=r, column=1).value: resumo.cell(row=r, column=2).value
            for r in range(1, 10)
        }
        assert labels.get("Registros extraidos") == 8
        assert labels.get("Paginas percorridas") == 2

        dados = wb["Dados"]
        assert dados.max_row == 9  # cabecalho + 8 registros

    def test_xlsx_sem_registros(self, tmp_path: Path) -> None:
        write_extract_artifacts([], _result([], total_pages=0), tmp_path / "saida")
        wb = load_workbook(tmp_path / "saida" / DATA_XLSX_NAME)
        dados = wb["Dados"]
        assert "Nenhum registro" in str(dados["A1"].value)
