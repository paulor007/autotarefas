"""Testes para autotarefas.tasks.send_artifacts (os 4 artefatos)."""

from __future__ import annotations

from typing import TYPE_CHECKING, Any

import httpx
import pandas as pd
import pytest
from openpyxl import load_workbook

from autotarefas.tasks.send_api import SendApiTask
from autotarefas.tasks.send_artifacts import (
    FAILED_CSV_NAME,
    IMPORT_JSON_NAME,
    RESULT_XLSX_NAME,
    SENT_CSV_NAME,
    write_send_artifacts,
)

if TYPE_CHECKING:
    from pathlib import Path

URL = "http://test.local/api/clientes"


def make_response(status_code: int, json_data: dict[str, Any] | None = None) -> httpx.Response:
    request = httpx.Request("POST", URL)
    return httpx.Response(status_code, json=json_data or {}, request=request)


@pytest.fixture
def cenario(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> tuple[pd.DataFrame, Any]:
    """
    Executa um envio real (mockado) com 3 destinos:
    linha 2 -> 201 (id 10) | linha 3 -> 422 | linha 4 -> 409.
    """
    csv = tmp_path / "leads.csv"
    pd.DataFrame(
        [
            {"nome": "Ana", "email": "ana@x.com", "cpf": "ok"},
            {"nome": "Bruno", "email": "bruno@x.com", "cpf": "invalido"},
            {"nome": "Carla", "email": "carla@x.com", "cpf": "duplicado"},
        ]
    ).to_csv(csv, index=False)

    def fake_post(url: str, **kwargs: Any) -> httpx.Response:
        payload = kwargs.get("json", {})
        if payload["cpf"] == "invalido":
            return make_response(422, {"error": "Validacao falhou", "detalhes": ["cpf"]})
        if payload["cpf"] == "duplicado":
            return make_response(409, {"error": "CPF ja cadastrado"})
        return make_response(201, {"status": "ok", "data": {"id": 10}})

    monkeypatch.setattr(httpx, "post", fake_post)
    task = SendApiTask(planilha_path=csv, url=URL)
    result = task.run()
    assert task.processed_dataframe is not None
    return task.processed_dataframe, result


class TestArtefatos:
    def test_gera_os_quatro_arquivos(
        self, cenario: tuple[pd.DataFrame, Any], tmp_path: Path
    ) -> None:
        df, result = cenario
        out = tmp_path / "saida"
        paths = write_send_artifacts(df, result, out)
        for p in paths:
            assert p.exists()
        nomes = sorted(p.name for p in paths)
        assert nomes == sorted([SENT_CSV_NAME, FAILED_CSV_NAME, RESULT_XLSX_NAME, IMPORT_JSON_NAME])

    def test_enviados_csv_com_id_externo(
        self, cenario: tuple[pd.DataFrame, Any], tmp_path: Path
    ) -> None:
        df, result = cenario
        out = tmp_path / "saida"
        write_send_artifacts(df, result, out)
        df = pd.read_csv(out / SENT_CSV_NAME)
        assert list(df["nome"]) == ["Ana"]
        assert int(df["_status_http"].iloc[0]) == 201
        assert str(df["_id_externo"].iloc[0]) == "10"

    def test_falhos_csv_com_motivo_e_categoria(
        self, cenario: tuple[pd.DataFrame, Any], tmp_path: Path
    ) -> None:
        df, result = cenario
        out = tmp_path / "saida"
        write_send_artifacts(df, result, out)
        df = pd.read_csv(out / FAILED_CSV_NAME)
        assert list(df["nome"]) == ["Bruno", "Carla"]
        assert list(df["_categoria"]) == ["validacao", "duplicado"]
        assert "Validacao falhou" in str(df["_motivo"].iloc[0])
        assert bool(df["_pode_reenviar"].iloc[0]) is False

    def test_xlsx_quatro_abas_e_resumo(
        self, cenario: tuple[pd.DataFrame, Any], tmp_path: Path
    ) -> None:
        df, result = cenario
        out = tmp_path / "saida"
        write_send_artifacts(df, result, out)
        wb = load_workbook(out / RESULT_XLSX_NAME)
        assert wb.sheetnames == ["Resumo", "Enviados", "Falhas", "Detalhes"]

        resumo = wb["Resumo"]
        labels = {
            resumo.cell(row=r, column=1).value: resumo.cell(row=r, column=2).value
            for r in range(1, 20)
        }
        assert labels.get("Total de registros") == 3
        assert labels.get("Enviados") == 1
        assert labels.get("Falhos") == 2
        # categorias com rotulo legivel
        assert "Dados invalidos (400/422)" in labels
        assert "Ja cadastrado (409)" in labels

        detalhes = wb["Detalhes"]
        header = [c.value for c in detalhes[1]]
        assert "idempotency_key" in header
        assert detalhes.max_row == 4  # cabecalho + 3 itens

    def test_reenvio_do_falhos_gera_as_mesmas_chaves(
        self, cenario: tuple[pd.DataFrame, Any], tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """A promessa do produto: registros_falhos.csv e reenviavel SEM duplicar."""
        df, result = cenario
        out = tmp_path / "saida"
        write_send_artifacts(df, result, out)

        # chaves originais das linhas que falharam (Bruno, Carla)
        originais = {i["idempotency_key"] for i in result.data["items"] if not i["sucesso"]}

        # reenvia o registros_falhos.csv como nova planilha de entrada
        keys_reenvio: list[str] = []

        def fake_post(url: str, **kwargs: Any) -> httpx.Response:
            keys_reenvio.append(kwargs["headers"]["Idempotency-Key"])
            return make_response(201, {"status": "ok", "data": {"id": 99}})

        monkeypatch.setattr(httpx, "post", fake_post)
        SendApiTask(planilha_path=out / FAILED_CSV_NAME, url=URL).run()

        # payload identico (colunas _ ignoradas) -> MESMAS chaves
        assert set(keys_reenvio) == originais
